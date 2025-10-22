"""
apps/products/import_export.py — Импорт/Экспорт товаров

Поддерживаемые форматы:
- CSV (разделитель: запятая или точка с запятой)
- JSON
- XML
- XLSX (Excel)

Использование:
    from apps.products.import_export import ProductImporter, ProductExporter
    
    # Импорт
    importer = ProductImporter(store)
    result = importer.import_from_file(file, file_format='csv')
    
    # Экспорт
    exporter = ProductExporter(store)
    file_content = exporter.export(products, format='xlsx')
"""

import csv
import json
import io
from decimal import Decimal, InvalidOperation
from typing import List, Dict, Any, Optional
from django.core.files.uploadedfile import UploadedFile
from defusedxml import ElementTree as ET
import openpyxl
from openpyxl.styles import Font, PatternFill
from .models import Product, Category, Size, ProductVariant


class ImportResult:
    """
    Результат импорта товаров.

    Содержит статистику: сколько создано, обновлено, ошибок.
    """

    def __init__(self):
        self.created = 0
        self.updated = 0
        self.errors = []
        self.total_rows = 0

    def add_error(self, row: int, message: str):
        """Добавить ошибку"""
        self.errors.append(f"Строка {row}: {message}")

    def to_dict(self):
        """Преобразовать в словарь для отображения"""
        return {
            'total_rows': self.total_rows,
            'created': self.created,
            'updated': self.updated,
            'errors': self.errors,
            'success': len(self.errors) == 0,
        }


class ProductImporter:
    """
    Класс для импорта товаров из файлов.

    Автоматически определяет формат файла по расширению.
    Поддерживает создание и обновление товаров.
    """

    # Обязательные поля для импорта
    REQUIRED_FIELDS = ['name', 'category_slug', 'retail_price']

    # Все доступные поля
    AVAILABLE_FIELDS = [
        'name', 'slug', 'description', 'short_description',
        'category_slug',  # Slug категории
        'retail_price', 'wholesale_price', 'discount_price',
        'stock', 'sku', 'available',
        'has_variants',  # Имеет ли варианты
        'specifications',  # JSON строка
        # Варианты (опционально):
        'variant_size', 'variant_stock', 'variant_sku',  # Один вариант
    ]

    def __init__(self, store):
        """
        Инициализация импортёра.

        Args:
            store: Магазин для которого импортируем товары
        """
        self.store = store

    def detect_format(self, file: UploadedFile) -> str:
        """
        Автоопределение формата файла по расширению.

        Args:
            file: Загруженный файл

        Returns:
            Формат файла: 'csv', 'json', 'xml', 'xlsx'

        Raises:
            ValueError: Если формат не поддерживается
        """
        filename = file.name.lower()

        if filename.endswith('.csv'):
            return 'csv'
        elif filename.endswith('.json'):
            return 'json'
        elif filename.endswith('.xml'):
            return 'xml'
        elif filename.endswith(('.xlsx', '.xls')):
            return 'xlsx'
        else:
            raise ValueError(f"Неподдерживаемый формат файла: {filename}")

    def import_from_file(self, file: UploadedFile, file_format: Optional[str] = None) -> ImportResult:
        """
        Импорт товаров из файла.

        Args:
            file: Загруженный файл
            file_format: Формат файла (если None - автоопределение)

        Returns:
            ImportResult с результатами импорта
        """
        # Автоопределение формата если не указан
        if not file_format:
            file_format = self.detect_format(file)

        # Парсинг файла в зависимости от формата
        if file_format == 'csv':
            rows = self._parse_csv(file)
        elif file_format == 'json':
            rows = self._parse_json(file)
        elif file_format == 'xml':
            rows = self._parse_xml(file)
        elif file_format == 'xlsx':
            rows = self._parse_xlsx(file)
        else:
            raise ValueError(f"Неподдерживаемый формат: {file_format}")

        # Импорт строк
        result = self._import_rows(rows)

        return result

    def _parse_csv(self, file: UploadedFile) -> List[Dict[str, Any]]:
        """
        Парсинг CSV файла.

        Поддерживает разделители: запятая (,) и точка с запятой (;)
        """
        content = file.read().decode('utf-8-sig')  # utf-8-sig для BOM

        # Определяем разделитель
        delimiter = ',' if ',' in content[:1000] else ';'

        reader = csv.DictReader(io.StringIO(content), delimiter=delimiter)
        return list(reader)

    def _parse_json(self, file: UploadedFile) -> List[Dict[str, Any]]:
        """Парсинг JSON файла"""
        content = file.read().decode('utf-8')
        data = json.loads(content)

        # Поддержка двух форматов:
        # 1. Массив объектов: [{...}, {...}]
        # 2. Объект с ключом products: {"products": [{...}]}
        if isinstance(data, list):
            return data
        elif isinstance(data, dict) and 'products' in data:
            return data['products']
        else:
            raise ValueError(
                "JSON должен быть массивом объектов или содержать ключ 'products'")

    def _parse_xml(self, file: UploadedFile) -> List[Dict[str, Any]]:
        """
        Парсинг XML файла.

        Формат:
        <products>
            <product>
                <name>Товар 1</name>
                <retail_price>1000</retail_price>
                ...
            </product>
        </products>
        """
        content = file.read()
        tree = ET.fromstring(content)

        rows = []
        for product_elem in tree.findall('product'):
            row = {}
            for child in product_elem:
                row[child.tag] = child.text
            rows.append(row)

        return rows

    def _parse_xlsx(self, file: UploadedFile) -> List[Dict[str, Any]]:
        """
        Парсинг XLSX (Excel) файла.

        Первая строка - заголовки.
        """
        workbook = openpyxl.load_workbook(file, data_only=True)
        sheet = workbook.active

        # Получаем заголовки из первой строки
        headers = []
        for cell in sheet[1]:
            headers.append(cell.value)

        # Читаем данные
        rows = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            row_dict = {}
            for header, value in zip(headers, row):
                if header and value is not None:
                    row_dict[header] = value
            if row_dict:  # Пропускаем пустые строки
                rows.append(row_dict)

        return rows

    def _import_rows(self, rows: List[Dict[str, Any]]) -> ImportResult:
        """
        Импорт строк данных в базу.

        Логика:
        - Если товар с таким SKU существует → обновление
        - Если нет → создание
        """
        result = ImportResult()
        result.total_rows = len(rows)

        # start=2 т.к. строка 1 = заголовки
        for index, row in enumerate(rows, start=2):
            try:
                # Валидация обязательных полей
                for field in self.REQUIRED_FIELDS:
                    if field not in row or not row[field]:
                        result.add_error(
                            index, f"Отсутствует обязательное поле: {field}")
                        continue

                # Получаем категорию
                category_slug = row.get('category_slug')
                try:
                    category = Category.objects.get(
                        store=self.store,
                        slug=category_slug
                    )
                except Category.DoesNotExist:
                    result.add_error(
                        index, f"Категория не найдена: {category_slug}")
                    continue

                # Преобразуем цены в Decimal
                try:
                    retail_price = Decimal(str(row['retail_price']))
                    wholesale_price = Decimal(str(row['wholesale_price'])) if row.get(
                        'wholesale_price') else None
                    discount_price = Decimal(str(row['discount_price'])) if row.get(
                        'discount_price') else None
                except (InvalidOperation, ValueError) as e:
                    result.add_error(index, f"Неверный формат цены: {e}")
                    continue

                # Проверяем существует ли товар (по SKU)
                sku = row.get('sku', '')
                if sku:
                    product = Product.objects.filter(
                        store=self.store,
                        sku=sku
                    ).first()
                else:
                    product = None

                # Подготовка данных
                product_data = {
                    'store': self.store,
                    'category': category,
                    'name': row['name'],
                    'slug': row.get('slug', ''),
                    'description': row.get('description', ''),
                    'short_description': row.get('short_description', ''),
                    'retail_price': retail_price,
                    'wholesale_price': wholesale_price,
                    'discount_price': discount_price,
                    'stock': int(row.get('stock', 0)),
                    'sku': sku,
                    'available': row.get('available', 'true').lower() in ('true', '1', 'да'),
                    'has_variants': row.get('has_variants', 'false').lower() in ('true', '1', 'да'),
                }

                # Specifications (JSON)
                if row.get('specifications'):
                    try:
                        product_data['specifications'] = json.loads(
                            row['specifications'])
                    except json.JSONDecodeError:
                        result.add_error(
                            index, "Неверный формат JSON в specifications")
                        continue

                # Создание или обновление товара
                if product:
                    # Обновление
                    for key, value in product_data.items():
                        setattr(product, key, value)
                    product.save()
                    result.updated += 1
                else:
                    # Создание
                    product = Product.objects.create(**product_data)
                    result.created += 1

                # Импорт варианта (если указан)
                if row.get('variant_size'):
                    self._import_variant(product, row, result, index)

            except Exception as e:
                result.add_error(index, f"Ошибка: {str(e)}")

        return result

    def _import_variant(self, product: Product, row: Dict[str, Any], result: ImportResult, row_index: int):
        """
        Импорт варианта товара (размера).

        Args:
            product: Товар
            row: Строка данных с полями variant_*
            result: Объект результата (для добавления ошибок)
            row_index: Номер строки (для ошибок)
        """
        try:
            # Получаем размер
            size_value = row['variant_size']
            size = Size.objects.filter(value=size_value).first()

            if not size:
                result.add_error(row_index, f"Размер не найден: {size_value}")
                return

            # Создаём или обновляем вариант
            variant, created = ProductVariant.objects.update_or_create(
                product=product,
                size=size,
                defaults={
                    'stock': int(row.get('variant_stock', 0)),
                    'sku': row.get('variant_sku', ''),
                    'is_active': True,
                }
            )

        except Exception as e:
            result.add_error(row_index, f"Ошибка импорта варианта: {e}")


class ProductExporter:
    """
    Класс для экспорта товаров в файлы.

    Поддерживает форматы: CSV, JSON, XML, XLSX.
    """

    def __init__(self, store):
        """
        Инициализация экспортёра.

        Args:
            store: Магазин для которого экспортируем товары
        """
        self.store = store

    def export(self, products, format='csv', include_variants=True) -> bytes:
        """
        Экспорт товаров в указанном формате.

        Args:
            products: QuerySet или список товаров
            format: Формат экспорта ('csv', 'json', 'xml', 'xlsx')
            include_variants: Включать ли варианты товаров

        Returns:
            bytes: Содержимое файла
        """
        if format == 'csv':
            return self._export_csv(products, include_variants)
        elif format == 'json':
            return self._export_json(products, include_variants)
        elif format == 'xml':
            return self._export_xml(products, include_variants)
        elif format == 'xlsx':
            return self._export_xlsx(products, include_variants)
        else:
            raise ValueError(f"Неподдерживаемый формат: {format}")

    def _export_csv(self, products, include_variants) -> bytes:
        """Экспорт в CSV"""
        output = io.StringIO()
        writer = csv.writer(output, delimiter=';')

        # Заголовки
        headers = [
            'name', 'slug', 'description', 'short_description',
            'category_slug', 'retail_price', 'wholesale_price', 'discount_price',
            'stock', 'sku', 'available', 'has_variants'
        ]

        if include_variants:
            headers.extend(['variant_size', 'variant_stock', 'variant_sku'])

        writer.writerow(headers)

        # Данные
        for product in products:
            if include_variants and product.has_variants:
                # Экспортируем товар с каждым вариантом отдельной строкой
                for variant in product.variants.filter(is_active=True):
                    row = self._product_to_row(product)
                    row.extend([
                        variant.size.value,
                        variant.stock,
                        variant.sku,
                    ])
                    writer.writerow(row)
            else:
                # Обычный товар
                row = self._product_to_row(product)
                if include_variants:
                    row.extend(['', '', ''])
                writer.writerow(row)

        return output.getvalue().encode('utf-8-sig')  # BOM для Excel

    def _export_json(self, products, include_variants) -> bytes:
        """Экспорт в JSON"""
        data = []

        for product in products:
            product_dict = {
                'name': product.name,
                'slug': product.slug,
                'description': product.description,
                'short_description': product.short_description,
                'category_slug': product.category.slug if product.category else '',
                'retail_price': str(product.retail_price),
                'wholesale_price': str(product.wholesale_price) if product.wholesale_price else None,
                'discount_price': str(product.discount_price) if product.discount_price else None,
                'stock': product.stock,
                'sku': product.sku,
                'available': product.available,
                'has_variants': product.has_variants,
            }

            if include_variants and product.has_variants:
                product_dict['variants'] = [
                    {
                        'size': v.size.value,
                        'stock': v.stock,
                        'sku': v.sku,
                    }
                    for v in product.variants.filter(is_active=True)
                ]

            data.append(product_dict)

        return json.dumps({'products': data}, ensure_ascii=False, indent=2).encode('utf-8')

    def _export_xml(self, products, include_variants) -> bytes:
        """Экспорт в XML"""
        root = ET.Element('products')

        for product in products:
            product_elem = ET.SubElement(root, 'product')

            ET.SubElement(product_elem, 'name').text = product.name
            ET.SubElement(product_elem, 'slug').text = product.slug
            ET.SubElement(
                product_elem, 'description').text = product.description or ''
            ET.SubElement(
                product_elem, 'category_slug').text = product.category.slug if product.category else ''
            ET.SubElement(product_elem, 'retail_price').text = str(
                product.retail_price)

            if product.wholesale_price:
                ET.SubElement(product_elem, 'wholesale_price').text = str(
                    product.wholesale_price)

            ET.SubElement(product_elem, 'stock').text = str(product.stock)
            ET.SubElement(product_elem, 'sku').text = product.sku
            ET.SubElement(product_elem, 'available').text = str(
                product.available).lower()

            if include_variants and product.has_variants:
                variants_elem = ET.SubElement(product_elem, 'variants')
                for variant in product.variants.filter(is_active=True):
                    var_elem = ET.SubElement(variants_elem, 'variant')
                    ET.SubElement(var_elem, 'size').text = variant.size.value
                    ET.SubElement(var_elem, 'stock').text = str(variant.stock)
                    ET.SubElement(var_elem, 'sku').text = variant.sku

        return ET.tostring(root, encoding='utf-8', xml_declaration=True)

    def _export_xlsx(self, products, include_variants) -> bytes:
        """Экспорт в XLSX (Excel)"""
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Products"

        # Стиль заголовков
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(
            start_color="366092", end_color="366092", fill_type="solid")

        # Заголовки
        headers = [
            'Название', 'Slug', 'Описание', 'Категория',
            'Цена розничная', 'Цена оптовая', 'Скидка',
            'Остаток', 'Артикул', 'Доступен', 'Варианты'
        ]

        if include_variants:
            headers.extend(['Размер', 'Остаток варианта', 'Артикул варианта'])

        for col, header in enumerate(headers, start=1):
            cell = sheet.cell(row=1, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill

        # Данные
        row_num = 2
        for product in products:
            if include_variants and product.has_variants:
                for variant in product.variants.filter(is_active=True):
                    self._write_product_row(sheet, row_num, product, variant)
                    row_num += 1
            else:
                self._write_product_row(sheet, row_num, product)
                row_num += 1

        # Автоширина колонок
        for column in sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            sheet.column_dimensions[column_letter].width = min(
                max_length + 2, 50)

        # Сохраняем в bytes
        output = io.BytesIO()
        workbook.save(output)
        return output.getvalue()

    def _product_to_row(self, product) -> list:
        """Преобразовать товар в строку CSV"""
        return [
            product.name,
            product.slug,
            product.description,
            product.short_description,
            product.category.slug if product.category else '',
            str(product.retail_price),
            str(product.wholesale_price) if product.wholesale_price else '',
            str(product.discount_price) if product.discount_price else '',
            product.stock,
            product.sku,
            'true' if product.available else 'false',
            'true' if product.has_variants else 'false',
        ]

    def _write_product_row(self, sheet, row_num, product, variant=None):
        """Записать строку товара в Excel"""
        data = [
            product.name,
            product.slug,
            product.description,
            product.category.slug if product.category else '',
            float(product.retail_price),
            float(product.wholesale_price) if product.wholesale_price else None,
            float(product.discount_price) if product.discount_price else None,
            product.stock,
            product.sku,
            'Да' if product.available else 'Нет',
            'Да' if product.has_variants else 'Нет',
        ]

        if variant:
            data.extend([
                variant.size.value,
                variant.stock,
                variant.sku,
            ])

        for col, value in enumerate(data, start=1):
            sheet.cell(row=row_num, column=col, value=value)
